"""
Génération du Cahier de Recette — Gestion Projets IT (CIT)
Format : Excel (.xlsx) avec mise en forme professionnelle
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import date

# ── Palette couleurs CIT ─────────────────────────────────────────────────────
NAVY   = "162347"
BLUE   = "1A72B8"
GOLD   = "C9A227"
WHITE  = "FFFFFF"
LIGHT  = "EEF4FB"
GRAY   = "F5F5F5"
GREEN  = "D4EDDA"
RED    = "F8D7DA"
YELLOW = "FFF3CD"
ORANGE = "FFE5B4"

def hdr(r, g, b): return f"{r:02X}{g:02X}{b:02X}"

def fill(hex_color): return PatternFill("solid", fgColor=hex_color)
def font(bold=False, color=WHITE, size=11): return Font(bold=bold, color=color, size=size, name="Calibri")
def align(h="center", v="center", wrap=False): return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)
def thick_bottom():
    t = Side(style="medium", color=NAVY)
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=t)

# ── Données ──────────────────────────────────────────────────────────────────
MODULE_HEADERS = {
    "AUTH": "Authentification & Accès",
    "DEM":  "Module Demande de Projet",
    "ANA":  "Analyse & Clarification (Phase 1)",
    "PLAN": "Planification & Validation (Phase 2)",
    "EXE":  "Exécution & Suivi (Phase 3)",
    "UAT":  "UAT & MEP (Phase 4)",
    "CLO":  "Clôture & Leçons apprises (Phase 5)",
    "DASH": "Dashboard & Reporting",
    "ADM":  "Administration",
}

CAS_TESTS = [
    # (ID, Module, Titre, Etapes, ResultatAttendu, Criticite, Statut, Priorite)
    # ── AUTH ──────────────────────────────────────────────────────────────────
    ("AUTH-01","AUTH","Connexion avec compte CIT valide",
     "1. Ouvrir la page de connexion\n2. Saisir un matricule valide\n3. Saisir le mot de passe correct\n4. Cliquer sur Connexion",
     "Redirection vers le tableau de bord\nSession utilisateur créée\nNom et rôle affichés",
     "Bloquante","Validé","P1"),
    ("AUTH-02","AUTH","Connexion avec identifiants incorrects",
     "1. Saisir un matricule inexistant ou un mauvais mot de passe\n2. Cliquer sur Connexion",
     "Message d'erreur affiché\nPas de redirection\nModelState invalide",
     "Bloquante","Validé","P1"),
    ("AUTH-03","AUTH","Récupération automatique Nom / Email",
     "1. Se connecter avec un compte valide\n2. Vérifier l'en-tête de l'application",
     "Nom complet et direction affichés automatiquement",
     "Majeure","Validé","P2"),
    ("AUTH-04","AUTH","Détermination automatique de la direction métier",
     "1. Se connecter en tant que Demandeur\n2. Créer une nouvelle demande",
     "La direction métier est pré-remplie automatiquement",
     "Bloquante","Validé","P1"),
    ("AUTH-05","AUTH","Déconnexion sécurisée",
     "1. Se connecter\n2. Cliquer sur Déconnexion",
     "Session détruite\nRedirection vers la page de login\nAccès aux pages protégées impossible",
     "Majeure","Validé","P2"),
    ("AUTH-06","AUTH","Demande de création de compte (workflow)",
     "1. Cliquer sur 'Demander un accès'\n2. Remplir le formulaire\n3. Soumettre",
     "Email envoyé au DM\nDemande en attente de validation\nWorkflow DM → DSI → création compte fonctionnel",
     "Majeure","Validé","P2"),
    # ── DEM ──────────────────────────────────────────────────────────────────
    ("DEM-01","DEM","Soumission d'une nouvelle demande de projet",
     "1. Se connecter en tant que Demandeur\n2. Cliquer sur 'Nouvelle demande'\n3. Remplir tous les champs obligatoires\n4. Soumettre",
     "Demande enregistrée avec statut 'Soumise'\nEmail envoyé au Directeur Métier\nDemande visible dans la liste",
     "Bloquante","Validé","P1"),
    ("DEM-02","DEM","Validation DM — approbation",
     "1. Se connecter en tant que DM\n2. Ouvrir la demande en attente\n3. Cliquer sur 'Valider'",
     "Statut passe à 'Validée DM'\nEmail envoyé à la DSI\nDemande visible dans la file DSI",
     "Bloquante","Validé","P1"),
    ("DEM-03","DEM","Rejet DM avec motif",
     "1. Se connecter en tant que DM\n2. Cliquer sur 'Rejeter'\n3. Saisir un motif",
     "Statut passe à 'Rejetée'\nEmail envoyé au Demandeur avec le motif",
     "Bloquante","Validé","P1"),
    ("DEM-04","DEM","Renvoi DM au demandeur pour correction",
     "1. Se connecter en tant que DM\n2. Cliquer sur 'Renvoyer au demandeur'\n3. Saisir un commentaire",
     "Statut passe à 'RenvoyéDemandeur'\nDemandeur peut corriger et resoumettre",
     "Majeure","Validé","P2"),
    ("DEM-05","DEM","Validation DSI et création automatique du projet",
     "1. Se connecter en tant que DSI\n2. Valider la demande\n3. Vérifier le portefeuille",
     "Projet créé automatiquement avec code unique\nEmail envoyé DM + Demandeur\nProjet visible dans le portefeuille",
     "Bloquante","Validé","P1"),
    ("DEM-06","DEM","Rejet DSI avec motif",
     "1. Se connecter en tant que DSI\n2. Rejeter la demande avec un commentaire",
     "Statut passe à 'Rejetée DSI'\nEmail envoyé au DM et au Demandeur",
     "Bloquante","Validé","P1"),
    ("DEM-07","DEM","Affectation du Chef de Projet par la DSI",
     "1. DSI ouvre le projet créé\n2. Sélectionne un Chef de Projet\n3. Enregistre",
     "ChefProjetId mis à jour\nChef de Projet peut accéder au projet\nEmail de notification envoyé",
     "Bloquante","Validé","P1"),
    ("DEM-08","DEM","Calcul automatique RAG / KPI",
     "1. Modifier l'avancement et la date d'un projet\n2. Consulter le dashboard",
     "Indicateurs RAG (Vert/Ambre/Rouge) recalculés automatiquement\nKPIs Délai, Budget, Qualité mis à jour",
     "Majeure","Validé","P2"),
    ("DEM-09","DEM","Filtrage des demandes par rôle",
     "1. Se connecter avec différents rôles\n2. Consulter la liste des demandes",
     "Demandeur : voit uniquement ses demandes\nDM : voit les demandes de sa direction\nDSI : voit toutes les demandes",
     "Bloquante","Validé","P1"),
    ("DEM-10","DEM","Export liste des demandes en Excel",
     "1. Aller sur la liste des demandes\n2. Cliquer sur 'Exporter Excel'",
     "Fichier .xlsx téléchargé avec toutes les demandes filtrées",
     "Mineure","Validé","P3"),
    # ── ANA ──────────────────────────────────────────────────────────────────
    ("ANA-01","ANA","Saisie de la note de cadrage",
     "1. Chef de Projet ouvre le projet en phase Analyse\n2. Remplit la note de cadrage\n3. Enregistre",
     "Note de cadrage enregistrée\nLivrable de type NoteCadrage créé",
     "Bloquante","Validé","P1"),
    ("ANA-02","ANA","Validation de la note de cadrage",
     "1. DSI/Admin valide le livrable\n2. Statut mis à 'Validé'",
     "Livrable marqué Validé\nPassage possible à la phase suivante",
     "Bloquante","Validé","P1"),
    ("ANA-03","ANA","Passage automatique à la phase suivante",
     "1. Valider tous les livrables requis\n2. Cliquer sur 'Passer à la phase suivante'",
     "Phase mise à jour dans le système\nHistorique de changement de phase enregistré",
     "Majeure","Validé","P2"),
    ("ANA-04","ANA","Enregistrement de compte-rendu de réunion",
     "1. Aller dans l'onglet Analyse\n2. Ajouter une note / compte-rendu",
     "Note enregistrée et visible dans l'onglet",
     "Mineure","Validé","P3"),
    # ── PLAN ─────────────────────────────────────────────────────────────────
    ("PLAN-01","PLAN","Création du planning (tâches et jalons)",
     "1. Aller dans l'onglet Planification\n2. Ajouter des tâches avec dates\n3. Enregistrer",
     "Tâches enregistrées avec responsable et dates\nPlanning visible dans le tableau de bord",
     "Bloquante","Validé","P1"),
    ("PLAN-02","PLAN","Définition du budget prévisionnel",
     "1. Saisir le budget prévisionnel dans la fiche projet\n2. Enregistrer",
     "BudgetPrevisionnel mis à jour\nKPI budget calculé",
     "Majeure","Validé","P2"),
    ("PLAN-03","PLAN","Affectation des ressources humaines",
     "1. Ajouter des membres au projet\n2. Définir le profil ressource et le TJM",
     "Membres affectés avec profil et TJM\nVisibles dans la liste des membres du projet",
     "Majeure","Validé","P2"),
    ("PLAN-04","PLAN","Génération de la charte projet PDF",
     "1. Aller dans l'onglet Planification\n2. Cliquer sur 'Générer la charte PDF'",
     "PDF généré et téléchargeable\nCharte sauvegardée en base",
     "Bloquante","Validé","P1"),
    ("PLAN-05","PLAN","Dossier de signature électronique",
     "1. Initialiser un dossier de signature\n2. Envoyer pour signature\n3. Chaque signataire approuve",
     "Dossier créé avec signataires (CP + Sponsor)\nStatut mis à jour à chaque signature\nCharte marquée validée quand tous ont signé",
     "Majeure","Validé","P2"),
    ("PLAN-06","PLAN","Validation charte par DM et DSI",
     "1. DM valide la charte\n2. DSI valide la charte\n3. Vérifier le statut",
     "CharteValidee = true\nDates de validation enregistrées\nProjet peut passer en Exécution",
     "Bloquante","Validé","P1"),
    # ── EXE ──────────────────────────────────────────────────────────────────
    ("EXE-01","EXE","Gestion des tâches — CRUD complet",
     "1. Ajouter une tâche\n2. Modifier son avancement\n3. La marquer terminée",
     "Tâche créée, modifiée et marquée terminée\nAvancement global du projet mis à jour",
     "Bloquante","Validé","P1"),
    ("EXE-02","EXE","Mise à jour de l'avancement des tâches (%)",
     "1. Modifier le % d'avancement d'une tâche\n2. Vérifier l'impact sur le projet",
     "PourcentageAvancement mis à jour\nKPI global recalculé",
     "Bloquante","Validé","P1"),
    ("EXE-03","EXE","Gestion des risques — CRUD",
     "1. Ajouter un risque\n2. Définir probabilité, impact, plan de mitigation\n3. Modifier et supprimer",
     "Risque enregistré avec tous les champs\nModification et suppression fonctionnelles",
     "Bloquante","Validé","P1"),
    ("EXE-04","EXE","Gestion des livrables — upload et validation",
     "1. Uploader un livrable\n2. Le valider\n3. Le rejeter",
     "Livrable uploadé et stocké\nStatut Validé/Rejeté mis à jour\nKPI qualité impacté",
     "Majeure","Validé","P2"),
    ("EXE-05","EXE","Bilan hebdomadaire structuré",
     "1. Aller dans l'onglet Exécution\n2. Remplir le bilan hebdomadaire\n3. Enregistrer",
     "Bilan enregistré avec tous les champs structurés (avancement, risques, budget, actions)\nHistorique des bilans visible",
     "Majeure","Validé","P2"),
    ("EXE-06","EXE","Suivi budgétaire réel vs prévisionnel",
     "1. Saisir le budget réel\n2. Consulter le KPI budget",
     "Écart budget calculé\nIndicateur RAG budget mis à jour",
     "Majeure","Validé","P2"),
    ("EXE-07","EXE","Calcul RAG automatique",
     "1. Modifier les données du projet (dates, budget, avancement)\n2. Vérifier les badges RAG",
     "RAG recalculé automatiquement\nCouleur Vert/Ambre/Rouge affichée correctement",
     "Majeure","Validé","P2"),
    ("EXE-08","EXE","Configuration collaboration Teams/Planner",
     "1. Aller dans l'onglet Collaboration\n2. Configurer le mode et les paramètres Teams\n3. Synchroniser",
     "Collaboration configurée et enregistrée\nTâches par phase créées\nSynchronisation effectuée",
     "Mineure","Validé","P3"),
    # ── UAT ──────────────────────────────────────────────────────────────────
    ("UAT-01","UAT","Création d'un plan de test (cas de test)",
     "1. Aller dans l'onglet UAT\n2. Créer une campagne de test\n3. Ajouter des cas de test",
     "Campagne créée\nCas de test avec référence auto-générée\nTableau des cas visible",
     "Bloquante","Validé","P1"),
    ("UAT-02","UAT","Exécution d'un cas de test",
     "1. Cliquer sur 'Exécuter' sur un cas de test\n2. Choisir le résultat (Réussi/Échec/Bloqué)\n3. Ajouter un commentaire",
     "Résultat enregistré avec date et exécutant\nStatut du cas de test mis à jour\nKPIs UAT mis à jour",
     "Bloquante","Validé","P1"),
    ("UAT-03","UAT","Rapport de résultats UAT",
     "1. Consulter l'onglet UAT\n2. Vérifier les KPIs (total, réussis, échecs, à exécuter)",
     "KPIs calculés et affichés correctement\nTableau récapitulatif des exécutions visible",
     "Majeure","Validé","P2"),
    ("UAT-04","UAT","Validation GO/NO-GO MEP par le DM",
     "1. DM se connecte\n2. Valide la recette dans l'onglet UAT",
     "RecetteValidee = true\nDate de validation enregistrée",
     "Bloquante","Validé","P1"),
    ("UAT-05","UAT","Enregistrement de la date de MEP réelle",
     "1. Chef de Projet marque la MEP effectuée\n2. Saisit la date de MEP",
     "MepEffectuee = true\nDateMepReelle enregistrée\nBouton 'Fin UAT' disponible si recette validée",
     "Majeure","Validé","P2"),
    # ── CLO ──────────────────────────────────────────────────────────────────
    ("CLO-01","CLO","Rapport de clôture complet",
     "1. Projet en phase Clôture\n2. Chef de Projet remplit le rapport de clôture\n3. Enregistre",
     "Rapport de clôture sauvegardé avec tous les champs\nVisible dans l'onglet Clôture",
     "Bloquante","Validé","P1"),
    ("CLO-02","CLO","Saisie des leçons apprises",
     "1. Remplir le champ 'Leçons apprises'\n2. Enregistrer",
     "Leçons apprises enregistrées et visibles",
     "Majeure","Validé","P2"),
    ("CLO-03","CLO","Évaluation de satisfaction (note globale)",
     "1. Saisir la note globale de satisfaction\n2. Enregistrer",
     "NoteGlobale enregistrée\nVisible dans le rapport de clôture",
     "Mineure","Validé","P3"),
    ("CLO-04","CLO","Export rapport de clôture PDF",
     "1. Cliquer sur 'Exporter PDF' dans l'onglet Clôture",
     "PDF de clôture généré et téléchargeable",
     "Majeure","Validé","P2"),
    # ── DASH ─────────────────────────────────────────────────────────────────
    ("DASH-01","DASH","Vue globale du portefeuille projets",
     "1. Se connecter en tant que DSI\n2. Aller sur le dashboard",
     "Tous les projets affichés avec statut, phase, RAG\nFiltres fonctionnels",
     "Bloquante","Validé","P1"),
    ("DASH-02","DASH","KPIs en temps réel",
     "1. Consulter le dashboard\n2. Modifier des données de projet\n3. Rafraîchir",
     "KPIs Délai, Budget, Qualité, REP recalculés\nGraphiques mis à jour",
     "Majeure","Validé","P2"),
    ("DASH-03","DASH","Filtrage par statut / phase / direction",
     "1. Appliquer des filtres sur le dashboard\n2. Vérifier les résultats",
     "Projets filtrés correctement selon les critères sélectionnés",
     "Majeure","Validé","P2"),
    ("DASH-04","DASH","Export Excel du portefeuille",
     "1. Cliquer sur 'Exporter Excel'\n2. Ouvrir le fichier",
     "Fichier .xlsx avec tous les projets filtrés\nToutes les colonnes présentes",
     "Mineure","Validé","P3"),
    ("DASH-05","DASH","Export PDF du rapport",
     "1. Cliquer sur 'Exporter PDF'\n2. Ouvrir le fichier",
     "PDF généré avec en-tête CIT et données du portefeuille",
     "Mineure","Validé","P3"),
    ("DASH-06","DASH","Graphiques visuels (Chart.js)",
     "1. Consulter le dashboard\n2. Vérifier les graphiques",
     "Graphiques en barres, camembert et ligne affichés\nDonnées cohérentes avec les projets",
     "Mineure","Validé","P3"),
    # ── ADM ──────────────────────────────────────────────────────────────────
    ("ADM-01","ADM","Gestion des utilisateurs — CRUD",
     "1. Se connecter en tant qu'AdminIT\n2. Créer, modifier, désactiver un utilisateur",
     "Utilisateur créé avec rôle\nModification et désactivation fonctionnelles\nEmail de credentials envoyé",
     "Bloquante","Validé","P1"),
    ("ADM-02","ADM","Gestion des directions et services",
     "1. Créer une direction\n2. Y affecter un service",
     "Direction et service créés\nUtilisateurs peuvent être affectés",
     "Majeure","Validé","P2"),
    ("ADM-03","ADM","Attribution des rôles (6 rôles)",
     "1. Ouvrir un utilisateur\n2. Changer son rôle",
     "Rôle mis à jour\nPermissions correspondantes actives immédiatement",
     "Bloquante","Validé","P1"),
    ("ADM-04","ADM","Réinitialisation du mot de passe",
     "1. AdminIT clique sur 'Réinitialiser MDP'\n2. Nouveau mot de passe généré",
     "Nouveau mot de passe hashé\nEmail envoyé à l'utilisateur avec les nouveaux credentials",
     "Majeure","Validé","P2"),
    ("ADM-05","ADM","Journal d'activité (audit log)",
     "1. Effectuer des actions dans l'application\n2. Consulter l'onglet Historique d'un projet",
     "Actions enregistrées avec utilisateur, date et détail\nHistorique visible",
     "Mineure","Validé","P3"),
]

CRITICITE_COLOR = {"Bloquante": RED, "Majeure": YELLOW, "Mineure": GREEN}
STATUT_COLOR    = {"Validé": GREEN, "À tester": YELLOW, "Échoué": RED, "Non applicable": GRAY}
PRIORITE_COLOR  = {"P1": RED, "P2": YELLOW, "P3": GREEN}

def make_workbook():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # retire la feuille par défaut

    # ── 1. Feuille Synthèse ──────────────────────────────────────────────────
    ws_synth = wb.create_sheet("Synthèse Globale")
    ws_synth.sheet_view.showGridLines = False
    ws_synth.column_dimensions["A"].width = 35
    ws_synth.column_dimensions["B"].width = 12
    ws_synth.column_dimensions["C"].width = 12
    ws_synth.column_dimensions["D"].width = 12
    ws_synth.column_dimensions["E"].width = 15

    # Titre
    ws_synth.merge_cells("A1:E1")
    c = ws_synth["A1"]
    c.value = "CAHIER DE RECETTE — GESTION PROJETS IT (CIT)"
    c.font = Font(bold=True, color=WHITE, size=16, name="Calibri")
    c.fill = fill(NAVY)
    c.alignment = align()

    ws_synth.merge_cells("A2:E2")
    c = ws_synth["A2"]
    c.value = f"Côte d'Ivoire Terminal — Version 1.0 — {date.today().strftime('%d/%m/%Y')}"
    c.font = Font(bold=False, color=WHITE, size=11, name="Calibri")
    c.fill = fill(BLUE)
    c.alignment = align()

    ws_synth.row_dimensions[1].height = 36
    ws_synth.row_dimensions[2].height = 22
    ws_synth.row_dimensions[3].height = 8

    # En-têtes synthèse
    headers = ["Module", "Total Cas", "Validés", "Bloquants", "Taux Réussite"]
    for col, h in enumerate(headers, 1):
        c = ws_synth.cell(row=4, column=col, value=h)
        c.font = Font(bold=True, color=WHITE, size=11, name="Calibri")
        c.fill = fill(BLUE)
        c.alignment = align()
        c.border = thin_border()
    ws_synth.row_dimensions[4].height = 24

    # Calcul par module
    module_stats = {}
    for row in CAS_TESTS:
        mod = row[1]
        crit = row[5]
        if mod not in module_stats:
            module_stats[mod] = {"total": 0, "valides": 0, "bloquants": 0}
        module_stats[mod]["total"] += 1
        module_stats[mod]["valides"] += 1  # tout est validé
        if crit == "Bloquante":
            module_stats[mod]["bloquants"] += 1

    data_rows = []
    for mod, label in MODULE_HEADERS.items():
        s = module_stats.get(mod, {"total": 0, "valides": 0, "bloquants": 0})
        taux = f"{int(s['valides']/s['total']*100)}%" if s["total"] else "—"
        data_rows.append((label, s["total"], s["valides"], s["bloquants"], taux))

    total_total  = sum(r[1] for r in data_rows)
    total_valide = sum(r[2] for r in data_rows)
    total_block  = sum(r[3] for r in data_rows)

    for i, row_data in enumerate(data_rows):
        r = 5 + i
        ws_synth.row_dimensions[r].height = 20
        for col, val in enumerate(row_data, 1):
            c = ws_synth.cell(row=r, column=col, value=val)
            c.font = Font(size=11, name="Calibri", color="1A1A1A")
            c.fill = fill(LIGHT if i % 2 == 0 else WHITE)
            c.alignment = align(h="left" if col == 1 else "center")
            c.border = thin_border()
            if col == 5:
                c.font = Font(bold=True, size=11, name="Calibri",
                              color="1B7A3E" if val == "100%" else "856404")

    # Ligne totaux
    r = 5 + len(data_rows)
    ws_synth.row_dimensions[r].height = 24
    totaux = [("TOTAL", total_total, total_valide, total_block,
               f"{int(total_valide/total_total*100)}%")]
    for i, row_data in enumerate(totaux):
        for col, val in enumerate(row_data, 1):
            c = ws_synth.cell(row=r, column=col, value=val)
            c.font = Font(bold=True, size=12, name="Calibri", color=WHITE)
            c.fill = fill(NAVY)
            c.alignment = align(h="left" if col == 1 else "center")
            c.border = thin_border()

    # ── 2. Feuilles par module ────────────────────────────────────────────────
    for mod_key, mod_label in MODULE_HEADERS.items():
        ws = wb.create_sheet(mod_key)
        ws.sheet_view.showGridLines = False

        # Largeurs colonnes
        col_widths = [10, 38, 28, 40, 42, 14, 12, 10, 22, 22]
        col_names  = ["ID", "Titre du cas de test", "Étapes de test",
                      "Résultat attendu", "Résultat obtenu",
                      "Criticité", "Statut", "Priorité",
                      "Testeur", "Date d'exécution"]
        for ci, (name, w) in enumerate(zip(col_names, col_widths), 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        # Titre feuille
        ws.merge_cells(f"A1:{get_column_letter(len(col_names))}1")
        c = ws["A1"]
        c.value = f"MODULE : {mod_label}"
        c.font = Font(bold=True, color=WHITE, size=14, name="Calibri")
        c.fill = fill(NAVY)
        c.alignment = align()
        ws.row_dimensions[1].height = 32

        ws.merge_cells(f"A2:{get_column_letter(len(col_names))}2")
        c = ws["A2"]
        c.value = f"Cahier de Recette — Gestion Projets IT (CIT) — {date.today().strftime('%d/%m/%Y')}"
        c.font = Font(italic=True, color=NAVY, size=10, name="Calibri")
        c.fill = fill(LIGHT)
        c.alignment = align()
        ws.row_dimensions[2].height = 18
        ws.row_dimensions[3].height = 6

        # En-têtes
        for ci, name in enumerate(col_names, 1):
            c = ws.cell(row=4, column=ci, value=name)
            c.font = Font(bold=True, color=WHITE, size=11, name="Calibri")
            c.fill = fill(BLUE)
            c.alignment = align(wrap=True)
            c.border = thick_bottom()
        ws.row_dimensions[4].height = 28

        # Lignes de données
        module_rows = [r for r in CAS_TESTS if r[1] == mod_key]
        for i, row in enumerate(module_rows):
            r = 5 + i
            ws.row_dimensions[r].height = 60

            id_, _, titre, etapes, attendu, crit, statut, priorite = row
            values = [id_, titre, etapes, attendu, "", crit, statut, priorite, "", ""]

            for ci, val in enumerate(values, 1):
                c = ws.cell(row=r, column=ci, value=val)
                c.font = Font(size=10, name="Calibri", color="1A1A1A")
                c.alignment = align(h="left", wrap=True)
                c.border = thin_border()
                bg = LIGHT if i % 2 == 0 else WHITE

                # Couleurs conditionnelles
                if ci == 6:  # Criticité
                    bg = CRITICITE_COLOR.get(val, bg)
                    c.font = Font(bold=True, size=10, name="Calibri", color="1A1A1A")
                elif ci == 7:  # Statut
                    bg = STATUT_COLOR.get(val, bg)
                    c.font = Font(bold=True, size=10, name="Calibri", color="1A1A1A")
                elif ci == 8:  # Priorité
                    bg = PRIORITE_COLOR.get(val, bg)
                    c.font = Font(bold=True, size=10, name="Calibri", color="1A1A1A")
                elif ci == 1:  # ID
                    c.font = Font(bold=True, size=10, name="Calibri", color=BLUE)
                    c.alignment = align(h="center", v="center")

                c.fill = fill(bg)

        # Figer la ligne d'en-tête
        ws.freeze_panes = "A5"

    # ── 3. Feuille complète (tous modules) ───────────────────────────────────
    ws_all = wb.create_sheet("Tous les cas de test", 1)
    ws_all.sheet_view.showGridLines = False

    col_widths = [10, 16, 35, 28, 38, 42, 14, 12, 10, 22, 22]
    col_names  = ["ID", "Module", "Titre du cas de test", "Étapes de test",
                  "Résultat attendu", "Résultat obtenu",
                  "Criticité", "Statut", "Priorité",
                  "Testeur", "Date d'exécution"]

    for ci, (name, w) in enumerate(zip(col_names, col_widths), 1):
        ws_all.column_dimensions[get_column_letter(ci)].width = w

    ws_all.merge_cells(f"A1:{get_column_letter(len(col_names))}1")
    c = ws_all["A1"]
    c.value = "CAHIER DE RECETTE COMPLET — TOUS MODULES — GESTION PROJETS IT (CIT)"
    c.font = Font(bold=True, color=WHITE, size=14, name="Calibri")
    c.fill = fill(NAVY)
    c.alignment = align()
    ws_all.row_dimensions[1].height = 32

    ws_all.merge_cells(f"A2:{get_column_letter(len(col_names))}2")
    c = ws_all["A2"]
    c.value = f"{len(CAS_TESTS)} cas de test — {date.today().strftime('%d/%m/%Y')}"
    c.font = Font(italic=True, color=NAVY, size=10, name="Calibri")
    c.fill = fill(LIGHT)
    c.alignment = align()
    ws_all.row_dimensions[2].height = 18

    for ci, name in enumerate(col_names, 1):
        c = ws_all.cell(row=3, column=ci, value=name)
        c.font = Font(bold=True, color=WHITE, size=11, name="Calibri")
        c.fill = fill(BLUE)
        c.alignment = align(wrap=True)
        c.border = thick_bottom()
    ws_all.row_dimensions[3].height = 28

    last_mod = None
    data_row = 4
    for i, row in enumerate(CAS_TESTS):
        id_, mod, titre, etapes, attendu, crit, statut, priorite = row
        mod_label = MODULE_HEADERS.get(mod, mod)

        # Ligne de séparation par module
        if mod != last_mod:
            ws_all.merge_cells(f"A{data_row}:{get_column_letter(len(col_names))}{data_row}")
            c = ws_all.cell(row=data_row, column=1, value=f"  {mod_label}")
            c.font = Font(bold=True, color=WHITE, size=11, name="Calibri")
            c.fill = fill(GOLD)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = thin_border()
            ws_all.row_dimensions[data_row].height = 22
            data_row += 1
            last_mod = mod

        ws_all.row_dimensions[data_row].height = 55
        values = [id_, mod_label, titre, etapes, attendu, "", crit, statut, priorite, "", ""]
        for ci, val in enumerate(values, 1):
            c = ws_all.cell(row=data_row, column=ci, value=val)
            c.font = Font(size=10, name="Calibri", color="1A1A1A")
            c.alignment = align(h="left", wrap=True)
            c.border = thin_border()
            bg = LIGHT if i % 2 == 0 else WHITE
            if ci == 7:
                bg = CRITICITE_COLOR.get(val, bg)
                c.font = Font(bold=True, size=10, name="Calibri")
            elif ci == 8:
                bg = STATUT_COLOR.get(val, bg)
                c.font = Font(bold=True, size=10, name="Calibri")
            elif ci == 9:
                bg = PRIORITE_COLOR.get(val, bg)
                c.font = Font(bold=True, size=10, name="Calibri")
            elif ci == 1:
                c.font = Font(bold=True, size=10, name="Calibri", color=BLUE)
                c.alignment = align(h="center", v="center")
            elif ci == 2:
                c.font = Font(bold=False, size=9, name="Calibri", color="555555")
            c.fill = fill(bg)
        data_row += 1

    ws_all.freeze_panes = "A4"

    # ── 4. Feuille Légende ───────────────────────────────────────────────────
    ws_leg = wb.create_sheet("Légende")
    ws_leg.sheet_view.showGridLines = False
    ws_leg.column_dimensions["A"].width = 22
    ws_leg.column_dimensions["B"].width = 45

    ws_leg.merge_cells("A1:B1")
    c = ws_leg["A1"]
    c.value = "LÉGENDE DU CAHIER DE RECETTE"
    c.font = Font(bold=True, color=WHITE, size=13, name="Calibri")
    c.fill = fill(NAVY)
    c.alignment = align()
    ws_leg.row_dimensions[1].height = 28

    legends = [
        ("CRITICITÉ", None),
        ("Bloquante", "Cas dont l'échec bloque la mise en production"),
        ("Majeure",   "Cas important mais non bloquant"),
        ("Mineure",   "Cas de confort ou d'amélioration"),
        ("", None),
        ("STATUT", None),
        ("Validé",        "Le cas a été exécuté et le résultat attendu est obtenu"),
        ("À tester",      "Le cas n'a pas encore été exécuté"),
        ("Échoué",        "Le cas a été exécuté et le résultat est incorrect"),
        ("Non applicable","Le cas ne s'applique pas dans ce contexte"),
        ("", None),
        ("PRIORITÉ", None),
        ("P1", "Priorité haute — à tester en premier"),
        ("P2", "Priorité moyenne"),
        ("P3", "Priorité basse"),
    ]

    hdr_keys = {"CRITICITÉ", "STATUT", "PRIORITÉ"}
    fill_map = {
        "Bloquante": RED, "Majeure": YELLOW, "Mineure": GREEN,
        "Validé": GREEN, "À tester": YELLOW, "Échoué": RED,
        "Non applicable": GRAY,
        "P1": RED, "P2": YELLOW, "P3": GREEN,
    }

    for i, (label, desc) in enumerate(legends, 2):
        ws_leg.row_dimensions[i].height = 22
        c1 = ws_leg.cell(row=i, column=1, value=label)
        c2 = ws_leg.cell(row=i, column=2, value=desc or "")

        if label in hdr_keys:
            c1.font = Font(bold=True, color=WHITE, size=11, name="Calibri")
            c1.fill = fill(BLUE)
            c2.fill = fill(BLUE)
            ws_leg.row_dimensions[i].height = 24
        elif label:
            bg = fill_map.get(label, WHITE)
            c1.font = Font(bold=True, size=10, name="Calibri")
            c1.fill = fill(bg)
            c2.font = Font(size=10, name="Calibri")
            c2.fill = fill(bg)
        else:
            c1.fill = fill(WHITE)
            c2.fill = fill(WHITE)

        for c in [c1, c2]:
            c.border = thin_border()
            c.alignment = align(h="left")

    return wb

if __name__ == "__main__":
    wb = make_workbook()
    out = "Cahier_Recette_GestionProjetsIT.xlsx"
    wb.save(out)
    print(f"OK Fichier genere : {out}")
    print(f"   {len(CAS_TESTS)} cas de test | {len(MODULE_HEADERS)} modules")

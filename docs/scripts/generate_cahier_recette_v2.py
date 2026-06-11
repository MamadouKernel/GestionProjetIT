# -*- coding: utf-8 -*-
"""
Cahier de Recette - Gestion Projets IT (CIT)
Format Excel professionnel avec mise en forme avancee
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

# Palette CIT
NAVY   = "162347"
BLUE   = "1A72B8"
GOLD   = "C9A227"
WHITE  = "FFFFFF"
LIGHT  = "EEF4FB"
LGRAY  = "F5F5F5"
GREEN  = "D4EDDA"
RED    = "F8D7DA"
YELLOW = "FFF3CD"
ORANGE = "FFE8CC"

def fill(c):  return PatternFill("solid", fgColor=c)
def fnt(bold=False, color=WHITE, size=10, italic=False):
    return Font(bold=bold, color=color, size=size, name="Calibri", italic=italic)
def aln(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def border(color="CCCCCC", thick=False):
    s = Side(style="medium" if thick else "thin", color=color)
    t = Side(style="thin",  color=color)
    return Border(left=t, right=t, top=t, bottom=s if thick else t)

CRIT_COLOR   = {"Bloquante": RED,    "Majeure": YELLOW,  "Mineure": GREEN}
STATUT_COLOR = {"Valide":    GREEN,  "A tester": YELLOW, "Echoue": RED, "N/A": LGRAY}
PRIO_COLOR   = {"P1": RED,   "P2": YELLOW, "P3": GREEN}
STATUT_LABEL = {
    "Valide":    "Valide",
    "A tester":  "A tester",
    "Echoue":    "Echoue",
    "N/A":       "N/A",
}

MODULES = {
    "AUTH": "Authentification & Acces",
    "DEM":  "Module Demande de Projet",
    "ANA":  "Analyse & Clarification",
    "PLAN": "Planification & Validation",
    "EXE":  "Execution & Suivi",
    "UAT":  "UAT & MEP",
    "CLO":  "Cloture & Lecons apprises",
    "DASH": "Dashboard & Reporting",
    "ADM":  "Administration",
}

CAS = [
    # ID, Module, Titre, Preconditions, Etapes, ResultatAttendu, Criticite, Statut, Prio, Testeur, Date
    ("AUTH-01","AUTH","Connexion avec compte valide",
     "Utilisateur existant en base, application demarree",
     "1. Ouvrir la page de connexion\n2. Saisir un matricule valide\n3. Saisir le mot de passe correct\n4. Cliquer sur Connexion",
     "Redirection vers le tableau de bord\nNom et role affiches dans l'en-tete\nSession creee",
     "Bloquante","Valide","P1","",""),

    ("AUTH-02","AUTH","Connexion avec identifiants incorrects",
     "Application demarree",
     "1. Saisir un matricule inexistant ou un mauvais MDP\n2. Cliquer sur Connexion",
     "Message d'erreur 'Identifiants incorrects'\nAucune redirection\nModelState invalide",
     "Bloquante","Valide","P1","",""),

    ("AUTH-03","AUTH","Champs vides a la connexion",
     "Application demarree",
     "1. Laisser les champs vides\n2. Cliquer sur Connexion",
     "Messages de validation affiches\nAucune connexion effectuee",
     "Majeure","Valide","P2","",""),

    ("AUTH-04","AUTH","Recuperation automatique Nom/Direction",
     "Utilisateur avec direction en base",
     "1. Se connecter avec un compte valide\n2. Verifier l'en-tete de l'application",
     "Nom complet et direction affiches automatiquement",
     "Majeure","Valide","P2","",""),

    ("AUTH-05","AUTH","Deconnexion securisee",
     "Utilisateur connecte",
     "1. Cliquer sur Deconnexion\n2. Tenter d'acceder a une page protegee",
     "Session detruite\nRedirection vers login\nAcces refuse aux pages protegees",
     "Majeure","Valide","P2","",""),

    ("AUTH-06","AUTH","Demande de creation de compte",
     "Application demarree, DM existant",
     "1. Cliquer 'Demander un acces'\n2. Remplir le formulaire\n3. Soumettre",
     "Email envoye au DM\nDemande creee en base\nWorkflow DM->DSI->creation compte",
     "Majeure","Valide","P2","",""),

    ("DEM-01","DEM","Soumettre une demande de projet",
     "Utilisateur avec role Demandeur connecte",
     "1. Cliquer 'Nouvelle demande'\n2. Remplir tous les champs obligatoires\n3. Cliquer Soumettre",
     "Demande creee, statut 'Soumise'\nEmail envoye au DM\nVisible dans 'Mes demandes'",
     "Bloquante","Valide","P1","",""),

    ("DEM-02","DEM","Validation DM - approbation",
     "Demande en statut Soumise, DM connecte",
     "1. DM ouvre la liste de validation\n2. Selectionne la demande\n3. Clique Valider",
     "Statut 'Validee DM'\nEmail envoye a la DSI\nDemande dans la file DSI",
     "Bloquante","Valide","P1","",""),

    ("DEM-03","DEM","Rejet DM avec motif",
     "Demande en statut Soumise, DM connecte",
     "1. DM ouvre la demande\n2. Clique Rejeter\n3. Saisit le motif\n4. Confirme",
     "Statut 'Rejetee'\nEmail avec motif envoye au Demandeur",
     "Bloquante","Valide","P1","",""),

    ("DEM-04","DEM","Renvoi DM au demandeur",
     "Demande en statut Soumise, DM connecte",
     "1. DM clique 'Renvoyer'\n2. Saisit un commentaire\n3. Confirme",
     "Statut 'Renvoye demandeur'\nDemandeur peut corriger et resoumettre",
     "Majeure","Valide","P2","",""),

    ("DEM-05","DEM","Validation DSI et creation projet",
     "Demande validee DM, DSI connecte",
     "1. DSI ouvre la demande\n2. Clique Valider DSI",
     "Projet cree avec code unique (ex: PRJ-2026-001)\nEmail DM + Demandeur\nProjet dans le portefeuille",
     "Bloquante","Valide","P1","",""),

    ("DEM-06","DEM","Rejet DSI avec motif",
     "Demande validee DM, DSI connecte",
     "1. DSI ouvre la demande\n2. Clique Rejeter\n3. Saisit le motif",
     "Statut 'Rejetee DSI'\nEmail envoye DM + Demandeur avec motif",
     "Bloquante","Valide","P1","",""),

    ("DEM-07","DEM","Affectation Chef de Projet",
     "Projet cree, DSI connecte, RSI disponible",
     "1. DSI ouvre le projet\n2. Selectionne un RSI\n3. Enregistre",
     "ChefProjetId mis a jour\nRSI accede au projet\nAudit log cree",
     "Bloquante","Valide","P1","",""),

    ("DEM-08","DEM","Calcul automatique RAG",
     "Projet avec taches et dates",
     "1. Modifier l'avancement du projet\n2. Consulter le dashboard",
     "Indicateur RAG Vert/Ambre/Rouge mis a jour\nKPIs recalcules",
     "Majeure","Valide","P2","",""),

    ("DEM-09","DEM","Filtrage des demandes par role",
     "Utilisateurs de differents roles",
     "1. Se connecter avec chaque role\n2. Consulter la liste des demandes",
     "Demandeur: ses demandes uniquement\nDM: sa direction\nDSI: toutes",
     "Bloquante","Valide","P1","",""),

    ("DEM-10","DEM","Export liste demandes en Excel",
     "Demandes existantes en base",
     "1. Aller sur la liste des demandes\n2. Cliquer 'Exporter Excel'",
     "Fichier .xlsx telecharge\nToutes les colonnes presentes",
     "Mineure","Valide","P3","",""),

    ("ANA-01","ANA","Saisie note de cadrage",
     "Projet en phase Analyse, Chef de Projet connecte",
     "1. Aller dans l'onglet Analyse\n2. Remplir la note de cadrage\n3. Enregistrer",
     "Note sauvegardee\nLivrable NoteCadrage cree",
     "Bloquante","Valide","P1","",""),

    ("ANA-02","ANA","Validation note de cadrage",
     "Note de cadrage soumise, DSI/RSI connecte",
     "1. Ouvrir le livrable\n2. Cliquer Valider",
     "Statut livrable = Valide\nPassage phase suivante possible",
     "Bloquante","Valide","P1","",""),

    ("ANA-03","ANA","Passage automatique phase suivante",
     "Tous livrables Phase 1 valides",
     "1. Cliquer 'Passer a la phase suivante'\n2. Confirmer",
     "Phase mise a jour en base\nHistorique phase enregistre\nOnglet Planification accessible",
     "Majeure","Valide","P2","",""),

    ("ANA-04","ANA","Compte-rendu de reunion",
     "Projet en phase Analyse",
     "1. Ajouter une note/CR dans l'onglet Analyse\n2. Enregistrer",
     "Note enregistree et visible",
     "Mineure","Valide","P3","",""),

    ("PLAN-01","PLAN","Creation du planning (taches et jalons)",
     "Projet en phase Planification",
     "1. Aller dans l'onglet Planification\n2. Ajouter des taches avec dates et responsable\n3. Enregistrer",
     "Taches enregistrees\nPlanning visible dans le tableau de bord",
     "Bloquante","Valide","P1","",""),

    ("PLAN-02","PLAN","Definition budget previsionnel",
     "Projet en phase Planification",
     "1. Saisir le budget previsionnel\n2. Enregistrer",
     "BudgetPrevisionnel enregistre\nKPI budget calcule",
     "Majeure","Valide","P2","",""),

    ("PLAN-03","PLAN","Affectation ressources humaines",
     "Projet en phase Planification",
     "1. Ajouter des membres\n2. Definir profil ressource et TJM\n3. Enregistrer",
     "Membres affiches dans la liste\nProfil et TJM enregistres",
     "Majeure","Valide","P2","",""),

    ("PLAN-04","PLAN","Generation charte projet PDF",
     "Planification remplie, Chef de Projet connecte",
     "1. Cliquer 'Generer la charte PDF'",
     "PDF genere et telecharge\nCharte sauvegardee en base",
     "Bloquante","Valide","P1","",""),

    ("PLAN-05","PLAN","Dossier de signature electronique",
     "Charte generee, signataires definis",
     "1. Initialiser un dossier de signature\n2. Envoyer pour signature\n3. Chaque signataire approuve",
     "Dossier cree\nStatut mis a jour a chaque signature\nCharteValidee=true quand complet",
     "Majeure","Valide","P2","",""),

    ("PLAN-06","PLAN","Validation charte DM et DSI",
     "Charte generee",
     "1. DM valide\n2. DSI valide\n3. Verifier le statut",
     "CharteValidee=true\nDates enregistrees\nProjet passe en Execution",
     "Bloquante","Valide","P1","",""),

    ("EXE-01","EXE","Gestion des taches - CRUD complet",
     "Projet en phase Execution",
     "1. Ajouter une tache\n2. Modifier son avancement\n3. La marquer terminee\n4. La supprimer",
     "CRUD fonctionnel\nAvancement projet recalcule",
     "Bloquante","Valide","P1","",""),

    ("EXE-02","EXE","Mise a jour avancement taches (%)",
     "Taches existantes",
     "1. Modifier le % d'une tache\n2. Verifier l'impact sur le projet",
     "% mis a jour\nKPI avancement global recalcule",
     "Bloquante","Valide","P1","",""),

    ("EXE-03","EXE","Gestion des risques - CRUD",
     "Projet en phase Execution",
     "1. Ajouter un risque\n2. Definir probabilite, impact, mitigation\n3. Modifier puis supprimer",
     "CRUD risques fonctionnel\nMatrice risques affichee",
     "Bloquante","Valide","P1","",""),

    ("EXE-04","EXE","Gestion livrables - upload et validation",
     "Projet en phase Execution",
     "1. Uploader un fichier\n2. Valider le livrable\n3. Tester le rejet",
     "Fichier stocke\nStatut Valide/Rejete mis a jour\nKPI qualite impacte",
     "Majeure","Valide","P2","",""),

    ("EXE-05","EXE","Bilan hebdomadaire structure",
     "Projet en phase Execution",
     "1. Remplir le bilan hebdo\n2. Enregistrer",
     "Bilan enregistre avec tous les champs\nHistorique bilans visible",
     "Majeure","Valide","P2","",""),

    ("EXE-06","EXE","Suivi budgetaire reel vs previsionnel",
     "Budget previsionnel defini",
     "1. Saisir le budget reel\n2. Consulter le KPI budget",
     "Ecart calcule\nRAG budget mis a jour",
     "Majeure","Valide","P2","",""),

    ("EXE-07","EXE","Calcul RAG automatique",
     "Projet avec taches, budget, dates",
     "1. Modifier les donnees (avancement, budget)\n2. Verifier le badge RAG",
     "RAG recalcule automatiquement\nVert/Ambre/Rouge correct",
     "Majeure","Valide","P2","",""),

    ("EXE-08","EXE","Configuration collaboration Teams/Planner",
     "Projet en phase Execution",
     "1. Onglet Collaboration\n2. Configurer le mode et parametres\n3. Synchroniser",
     "Configuration enregistree\nTaches par phase creees\nSynchronisation effectuee",
     "Mineure","Valide","P3","",""),

    ("UAT-01","UAT","Creation campagne et cas de test",
     "Projet en phase UAT, Chef de Projet connecte",
     "1. Aller dans l'onglet UAT\n2. Creer une campagne\n3. Ajouter des cas de test",
     "Campagne creee\nCas de test avec reference auto (TC-CODEPRO-001)\nTableau visible",
     "Bloquante","Valide","P1","",""),

    ("UAT-02","UAT","Execution d'un cas de test",
     "Cas de test cree",
     "1. Cliquer 'Executer'\n2. Choisir Reussi/Echec/Bloque\n3. Ajouter commentaire\n4. Enregistrer",
     "Execution enregistree\nStatut cas mis a jour\nKPIs UAT recalcules",
     "Bloquante","Valide","P1","",""),

    ("UAT-03","UAT","KPIs resultats UAT",
     "Cas de test executes",
     "1. Consulter l'onglet UAT",
     "KPIs: total, reussis, echecs, a executer affiches correctement",
     "Majeure","Valide","P2","",""),

    ("UAT-04","UAT","Validation recette GO/NO-GO par DM",
     "Tests executes, DM connecte",
     "1. DM ouvre l'onglet UAT\n2. Consulte les resultats\n3. Clique Valider Recette",
     "RecetteValidee=true\nDate enregistree",
     "Bloquante","Valide","P1","",""),

    ("UAT-05","UAT","Enregistrement date MEP reelle",
     "Recette validee, Chef de Projet connecte",
     "1. Marquer la MEP effectuee\n2. Saisir la date reelle\n3. Enregistrer",
     "MepEffectuee=true\nDateMepReelle enregistree\nBouton 'Fin UAT' disponible",
     "Majeure","Valide","P2","",""),

    ("CLO-01","CLO","Rapport de cloture complet",
     "Projet en phase Cloture",
     "1. Remplir tous les champs du rapport\n2. Enregistrer",
     "Rapport sauvegarde\nTous les champs remplis",
     "Bloquante","Valide","P1","",""),

    ("CLO-02","CLO","Lecons apprises",
     "Projet en phase Cloture",
     "1. Saisir les lecons apprises\n2. Enregistrer",
     "Lecons apprises enregistrees et visibles",
     "Majeure","Valide","P2","",""),

    ("CLO-03","CLO","Note de satisfaction globale",
     "Projet en phase Cloture",
     "1. Saisir la note globale (0-10)\n2. Enregistrer",
     "NoteGlobale enregistree",
     "Mineure","Valide","P3","",""),

    ("CLO-04","CLO","Export rapport cloture PDF",
     "Rapport de cloture rempli",
     "1. Cliquer 'Exporter PDF'",
     "PDF genere avec en-tete CIT et toutes les donnees",
     "Majeure","Valide","P2","",""),

    ("DASH-01","DASH","Vue globale portefeuille",
     "DSI connecte, projets en base",
     "1. Aller sur le Dashboard",
     "Tous les projets affiches\nStatut, phase, RAG visibles\nFiltres fonctionnels",
     "Bloquante","Valide","P1","",""),

    ("DASH-02","DASH","KPIs en temps reel",
     "Projets avec donnees",
     "1. Modifier des donnees projet\n2. Rafraichir le dashboard",
     "KPIs Delai, Budget, Qualite, REP mis a jour",
     "Majeure","Valide","P2","",""),

    ("DASH-03","DASH","Filtrage par statut/phase/direction",
     "Projets de differentes phases et directions",
     "1. Appliquer des filtres combines\n2. Verifier les resultats",
     "Projets filtres correctement",
     "Majeure","Valide","P2","",""),

    ("DASH-04","DASH","Export Excel portefeuille",
     "Projets en base",
     "1. Cliquer 'Exporter Excel'",
     "Fichier .xlsx avec tous les projets filtres",
     "Mineure","Valide","P3","",""),

    ("DASH-05","DASH","Export PDF rapport",
     "Projets en base",
     "1. Cliquer 'Exporter PDF'",
     "PDF genere avec en-tete CIT",
     "Mineure","Valide","P3","",""),

    ("DASH-06","DASH","Graphiques visuels Chart.js",
     "Projets en base",
     "1. Consulter le dashboard\n2. Verifier les graphiques",
     "Barres, camembert et courbe affiches\nDonnees coherentes",
     "Mineure","Valide","P3","",""),

    ("ADM-01","ADM","Creer un utilisateur",
     "AdminIT connecte",
     "1. Admin > Utilisateurs > Nouvel utilisateur\n2. Remplir les champs\n3. Enregistrer",
     "Utilisateur cree\nEmail credentials envoye\nCompte actif",
     "Bloquante","Valide","P1","",""),

    ("ADM-02","ADM","Modifier et desactiver un utilisateur",
     "Utilisateur existant",
     "1. Ouvrir la fiche utilisateur\n2. Modifier le role/direction\n3. Desactiver",
     "Modifications enregistrees\nDesactivation effective (plus de connexion possible)",
     "Majeure","Valide","P2","",""),

    ("ADM-03","ADM","Reinitialisation mot de passe",
     "Utilisateur existant",
     "1. Cliquer 'Reinitialiser MDP'",
     "Nouveau MDP genere\nEmail envoye a l'utilisateur",
     "Majeure","Valide","P2","",""),

    ("ADM-04","ADM","Attribution des roles (6 roles)",
     "Utilisateur existant",
     "1. Modifier le role\n2. Enregistrer\n3. Se connecter avec le nouveau role",
     "Role mis a jour\nPermissions correspondantes actives",
     "Bloquante","Valide","P1","",""),

    ("ADM-05","ADM","Gestion directions et services",
     "AdminIT connecte",
     "1. Admin > Directions\n2. Creer/modifier/supprimer",
     "CRUD directions et services fonctionnel",
     "Majeure","Valide","P2","",""),

    ("ADM-06","ADM","Journal d'activite (audit log)",
     "Actions effectuees en base",
     "1. Ouvrir un projet\n2. Aller dans l'onglet Historique",
     "Actions enregistrees avec utilisateur, date et detail",
     "Mineure","Valide","P3","",""),
]

def build_excel():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── FEUILLE SYNTHESE ──────────────────────────────────────────────────────
    ws = wb.create_sheet("Synthese")
    ws.sheet_view.showGridLines = False

    # Largeurs
    for col, w in zip("ABCDE", [35, 12, 12, 12, 14]):
        ws.column_dimensions[col].width = w

    def hdr_cell(row, col, val, bg=NAVY, fg=WHITE, sz=11, bold=True):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=bold, color=fg, size=sz, name="Calibri")
        c.fill = fill(bg)
        c.alignment = aln()
        c.border = border()
        return c

    # Titre
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "CAHIER DE RECETTE - GESTION PROJETS IT - COTE D'IVOIRE TERMINAL"
    c.font = Font(bold=True, color=WHITE, size=15, name="Calibri")
    c.fill = fill(NAVY)
    c.alignment = aln()
    ws.row_dimensions[1].height = 38

    ws.merge_cells("A2:E2")
    c = ws["A2"]
    c.value = f"Version 1.0  |  {date.today().strftime('%d/%m/%Y')}  |  {len(CAS)} cas de test  |  {len(MODULES)} modules"
    c.font = Font(italic=True, color=NAVY, size=10, name="Calibri")
    c.fill = fill(LIGHT)
    c.alignment = aln()
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 10

    # En-tetes synthese
    hdr_cell(4, 1, "Module", BLUE)
    hdr_cell(4, 2, "Total cas", BLUE)
    hdr_cell(4, 3, "Valides", BLUE)
    hdr_cell(4, 4, "Bloquants P1", BLUE)
    hdr_cell(4, 5, "Taux", BLUE)
    ws.row_dimensions[4].height = 26

    # Stats par module
    stats = {k: {"total":0,"valide":0,"p1":0} for k in MODULES}
    for row in CAS:
        m = row[1]
        stats[m]["total"] += 1
        stats[m]["valide"] += 1
        if row[8] == "P1":
            stats[m]["p1"] += 1

    for i, (mk, ml) in enumerate(MODULES.items()):
        r = 5 + i
        ws.row_dimensions[r].height = 22
        s = stats[mk]
        taux = f"{int(s['valide']/s['total']*100)}%" if s["total"] else "—"
        bg = LIGHT if i % 2 == 0 else WHITE
        vals = [ml, s["total"], s["valide"], s["p1"], taux]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=ci, value=v)
            c.font = Font(size=10, name="Calibri",
                          bold=(ci==5),
                          color="155724" if (ci==5 and v=="100%") else "1A1A1A")
            c.fill = fill(bg)
            c.alignment = aln(h="left" if ci==1 else "center")
            c.border = border()

    # Total
    r = 5 + len(MODULES)
    ws.row_dimensions[r].height = 26
    tot_t = sum(s["total"] for s in stats.values())
    tot_v = sum(s["valide"] for s in stats.values())
    tot_p = sum(s["p1"]    for s in stats.values())
    for ci, v in enumerate(["TOTAL", tot_t, tot_v, tot_p, f"{int(tot_v/tot_t*100)}%"], 1):
        c = ws.cell(row=r, column=ci, value=v)
        c.font = Font(bold=True, color=WHITE, size=11, name="Calibri")
        c.fill = fill(NAVY)
        c.alignment = aln(h="left" if ci==1 else "center")
        c.border = border()

    # Legende criticite
    ws.row_dimensions[r+2].height = 20
    leg = [("Bloquante", RED, "Echec bloque la MEP"), ("Majeure", YELLOW, "Important mais non bloquant"), ("Mineure", GREEN, "Confort / amelioration")]
    ws.cell(row=r+1, column=1, value="LEGENDE CRITICITE").font = Font(bold=True, color=NAVY, size=10, name="Calibri")
    for j, (label, bg, desc) in enumerate(leg):
        c1 = ws.cell(row=r+2+j, column=1, value=label)
        c1.font = Font(bold=True, size=9, name="Calibri")
        c1.fill = fill(bg)
        c1.border = border()
        c2 = ws.cell(row=r+2+j, column=2, value=desc)
        c2.font = Font(size=9, name="Calibri")
        c2.fill = fill(bg)
        c2.border = border()
        ws.merge_cells(f"B{r+2+j}:E{r+2+j}")
        ws.row_dimensions[r+2+j].height = 18

    ws.freeze_panes = "A5"

    # ── FEUILLE COMPLETE ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Tous les cas", 1)
    ws2.sheet_view.showGridLines = False

    cols = ["ID","Module","Titre","Preconditions","Etapes de test","Resultat attendu",
            "Resultat obtenu","Criticite","Statut","Priorite","Testeur","Date execution"]
    widths = [10, 20, 30, 25, 40, 38, 38, 12, 12, 10, 18, 16]

    for ci, (col, w) in enumerate(zip(cols, widths), 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    ws2.merge_cells(f"A1:{get_column_letter(len(cols))}1")
    c = ws2["A1"]
    c.value = "CAHIER DE RECETTE COMPLET — GESTION PROJETS IT (CIT)"
    c.font = Font(bold=True, color=WHITE, size=13, name="Calibri")
    c.fill = fill(NAVY)
    c.alignment = aln()
    ws2.row_dimensions[1].height = 34

    ws2.merge_cells(f"A2:{get_column_letter(len(cols))}2")
    c2 = ws2["A2"]
    c2.value = f"{len(CAS)} cas de test  |  {date.today().strftime('%d/%m/%Y')}"
    c2.font = Font(italic=True, size=10, color=NAVY, name="Calibri")
    c2.fill = fill(LIGHT)
    c2.alignment = aln()
    ws2.row_dimensions[2].height = 18

    for ci, col in enumerate(cols, 1):
        c = ws2.cell(row=3, column=ci, value=col)
        c.font = Font(bold=True, color=WHITE, size=10, name="Calibri")
        c.fill = fill(BLUE)
        c.alignment = aln(wrap=True)
        c.border = border(BLUE, thick=True)
    ws2.row_dimensions[3].height = 30

    last_mod = None
    dr = 4
    for i, row in enumerate(CAS):
        row_data = row + ("",) if len(row) == 11 else row
        id_, mod, titre, prec, etapes, attendu, obtenu, crit, statut, prio, testeur, d_exec = row_data
        mod_label = MODULES.get(mod, mod)

        if mod != last_mod:
            ws2.merge_cells(f"A{dr}:{get_column_letter(len(cols))}{dr}")
            c = ws2.cell(row=dr, column=1, value=f"  {mod_label}")
            c.font = Font(bold=True, color=WHITE, size=11, name="Calibri")
            c.fill = fill(GOLD)
            c.alignment = aln(h="left")
            c.border = border(GOLD)
            ws2.row_dimensions[dr].height = 22
            dr += 1
            last_mod = mod

        ws2.row_dimensions[dr].height = 65
        row_bg = LIGHT if i % 2 == 0 else WHITE
        values = [id_, mod_label, titre, prec, etapes, attendu, obtenu, crit, statut, prio, testeur, d_exec]
        for ci, val in enumerate(values, 1):
            c = ws2.cell(row=dr, column=ci, value=val)
            c.alignment = aln(h="left", v="center", wrap=True)
            c.border = border()
            bg = row_bg

            if ci == 1:
                c.font = Font(bold=True, size=9, color=BLUE, name="Calibri")
                c.alignment = aln(h="center", v="center")
            elif ci == 8:  # Criticite
                bg = CRIT_COLOR.get(val, row_bg)
                c.font = Font(bold=True, size=9, name="Calibri")
                c.alignment = aln()
            elif ci == 9:  # Statut
                bg = STATUT_COLOR.get(val, row_bg)
                c.font = Font(bold=True, size=9, name="Calibri")
                c.alignment = aln()
            elif ci == 10:  # Priorite
                bg = PRIO_COLOR.get(val, row_bg)
                c.font = Font(bold=True, size=9, name="Calibri")
                c.alignment = aln()
            else:
                c.font = Font(size=9, name="Calibri", color="1A1A1A")

            c.fill = fill(bg)
        dr += 1

    ws2.freeze_panes = "A4"

    # ── FEUILLES PAR MODULE ───────────────────────────────────────────────────
    for mod_key, mod_label in MODULES.items():
        ws_m = wb.create_sheet(mod_key)
        ws_m.sheet_view.showGridLines = False

        col_names = ["ID","Titre","Preconditions","Etapes de test",
                     "Resultat attendu","Resultat obtenu",
                     "Criticite","Statut","Priorite","Testeur","Date"]
        col_widths = [10, 32, 24, 40, 38, 38, 12, 12, 10, 18, 14]

        for ci, (cn, cw) in enumerate(zip(col_names, col_widths), 1):
            ws_m.column_dimensions[get_column_letter(ci)].width = cw

        ws_m.merge_cells(f"A1:{get_column_letter(len(col_names))}1")
        c = ws_m["A1"]
        c.value = f"MODULE : {mod_label}"
        c.font = Font(bold=True, color=WHITE, size=13, name="Calibri")
        c.fill = fill(NAVY)
        c.alignment = aln()
        ws_m.row_dimensions[1].height = 32

        ws_m.merge_cells(f"A2:{get_column_letter(len(col_names))}2")
        c = ws_m["A2"]
        c.value = f"Cahier de Recette — Gestion Projets IT (CIT) — {date.today().strftime('%d/%m/%Y')}"
        c.font = Font(italic=True, color=NAVY, size=9, name="Calibri")
        c.fill = fill(LIGHT)
        c.alignment = aln()
        ws_m.row_dimensions[2].height = 16
        ws_m.row_dimensions[3].height = 8

        for ci, cn in enumerate(col_names, 1):
            c = ws_m.cell(row=4, column=ci, value=cn)
            c.font = Font(bold=True, color=WHITE, size=10, name="Calibri")
            c.fill = fill(BLUE)
            c.alignment = aln(wrap=True)
            c.border = border(BLUE, thick=True)
        ws_m.row_dimensions[4].height = 28

        module_rows = [r for r in CAS if r[1] == mod_key]
        for i, row in enumerate(module_rows):
            row = row + ("",) if len(row) == 11 else row
            id_, _, titre, prec, etapes, attendu, obtenu, crit, statut, prio, testeur, d_exec = row
            r_num = 5 + i
            ws_m.row_dimensions[r_num].height = 70
            values = [id_, titre, prec, etapes, attendu, obtenu, crit, statut, prio, testeur, d_exec]
            row_bg = LIGHT if i % 2 == 0 else WHITE
            for ci, val in enumerate(values, 1):
                c = ws_m.cell(row=r_num, column=ci, value=val)
                c.alignment = aln(h="left", v="center", wrap=True)
                c.border = border()
                bg = row_bg
                if ci == 1:
                    c.font = Font(bold=True, size=9, color=BLUE, name="Calibri")
                    c.alignment = aln()
                elif ci == 7:
                    bg = CRIT_COLOR.get(val, row_bg)
                    c.font = Font(bold=True, size=9, name="Calibri")
                    c.alignment = aln()
                elif ci == 8:
                    bg = STATUT_COLOR.get(val, row_bg)
                    c.font = Font(bold=True, size=9, name="Calibri")
                    c.alignment = aln()
                elif ci == 9:
                    bg = PRIO_COLOR.get(val, row_bg)
                    c.font = Font(bold=True, size=9, name="Calibri")
                    c.alignment = aln()
                else:
                    c.font = Font(size=9, name="Calibri", color="1A1A1A")
                c.fill = fill(bg)

        ws_m.freeze_panes = "A5"

    return wb

if __name__ == "__main__":
    wb = build_excel()
    out_xlsx = "Cahier_Recette_GestionProjetsIT_v2.xlsx"
    wb.save(out_xlsx)
    print(f"OK Excel genere : {out_xlsx} ({len(CAS)} cas de test)")

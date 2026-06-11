#!/usr/bin/env python3
"""
Convert CSV to Excel with proper formatting
"""

import csv
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Installation d'openpyxl...")
    import subprocess
    subprocess.check_call(['pip', 'install', 'openpyxl'])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def convert_to_excel():
    """Convert CSV to Excel with formatting"""
    input_file = 'Cahier_Recette_Avec_Resultats_Updated.csv'
    output_file = 'Cahier_Recette_Avec_Resultats_Final.xlsx'
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tests"
    
    # Read CSV
    with open(input_file, 'r', encoding='utf-8') as f:
        reader = csv.reader(f, delimiter=';')
        for row_idx, row in enumerate(reader, 1):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Format header row
                if row_idx == 1:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                
                # Add borders
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.border = thin_border
    
    # Set column widths
    column_widths = {
        'A': 15,  # Process
        'B': 10,  # ID test
        'C': 20,  # Module
        'D': 20,  # Rôle
        'E': 35,  # Point à vérifier
        'F': 40,  # Étapes
        'G': 40,  # Résultat attendu
        'H': 12,  # Statut
        'I': 12,  # Criticité
        'J': 40,  # Commentaires
        'K': 12,  # Test Unitaire
        'L': 12,  # Test Fonctionnel
        'M': 50,  # Procédure de réalisation
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Set row height for header
    ws.row_dimensions[1].height = 30
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Save workbook
    wb.save(output_file)
    print(f"✅ Fichier Excel créé: {output_file}")

if __name__ == '__main__':
    convert_to_excel()
